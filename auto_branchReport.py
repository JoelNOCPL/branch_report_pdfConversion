#!python3

import openpyxl, os, shutil, requests, time, PIL,fpdf
from openpyxl import load_workbook
from fpdf import FPDF
from datetime import datetime






# ['Audit Parameters', 'Question Type', 'Checkpoints', 'Parameters', 'Points', 'Points Scored', "Auditor's Remarks", 'Deviation Closure Remarks', 'Closure Evidene']

def add_headers(pdf,headerPos):
    pdf.set_font(family="Arial",style='B',size=10)
    pdf.set_xy(10,headerPos)
    pdf.multi_cell(w=20,h=8,txt = 'Sr. \nNo',border=1,align='C',fill=True)
    pdf.set_xy(30,headerPos)
    pdf.multi_cell(w=30,h=8,txt = 'Audit \nParameters',border=1,align='C',fill=True)
    pdf.set_xy(60,headerPos)
    pdf.multi_cell(w=20,h=8,txt = 'Question \nType',border=1,align='C',fill=True)
    pdf.set_xy(80,headerPos)
    pdf.multi_cell(w=50,h=16,txt = 'Checkpoints',border=1,align='C',fill=True)
    pdf.set_xy(130,headerPos)
    pdf.multi_cell(w=70,h=16,txt = 'Parameters',border=1,align='C',fill=True)
    pdf.set_xy(200,headerPos)
    pdf.multi_cell(w=20,h=16,txt = 'Points',border=1,align='C',fill=True)
    pdf.set_xy(220,headerPos)
    pdf.multi_cell(w=20,h=8,txt = 'Points \nScored',border=1,align='C',fill=True)
    pdf.set_xy(240,headerPos)
    pdf.multi_cell(w=50,h=8,txt = "Auditor's \nRemarks",border=1,align='C',fill=True)
    pdf.set_xy(290,headerPos)
    pdf.multi_cell(w=40,h=8,txt = 'Deviation Closure \nRemarks',border=1,align='C',fill=True)
    pdf.set_xy(330,headerPos)
    pdf.multi_cell(w=70,h=16,txt = 'Closure Evidence',border=1,align='C',fill=True)
    pdf.set_font(family="Arial",style='',size=10)






def excel2pdf(excelPath):
    #os.chdir(r'C:\\Users\\Joel\\NOCPL\\BranchReport_pdfs')
    wb = load_workbook(excelPath)

    sheetName = wb[wb.get_sheet_names()[0]]
    ws = wb[wb.sheetnames[0]]
    k = ws.max_row

    columns = ['B','E','G','I','J','K','L','M','N']


    headers = [ws[a+str(18)].value for a in columns]
    #print(headers)

    widths = [    20,    30,   20,   50,    70,   20,     20,     50,   40,   70]
    #[10,50,70,120,190,205,220,270,310]

    xCoords = [10,  30,     60,   80,   130,   200,   220,    240,   290,   330 ]


    #pdf = FPDF()
    pdf = FPDF('L', 'mm', (297, 410)) # landscape mode
    pdf.add_page()

    pdf.set_text_color(0,0,0)
    #pdf.set_fill_color(r: int, g: int = -1, b: int = -1)
    pdf.set_fill_color(255,255,0)

    imageFolder = '.\\imageDownloads\\'

    
    description_data=[]
    
    
    ##### for description data-table ######
    j=0
    r=4
    while True:
        if j>=10:
            break
        column_data = [ws[a+str(r)].value for a in ['B','C','D','E','F']] # get all column values
        column_data = [t for t in column_data if t!=None] # keep only column values with not-NULL data
        if column_data==[] or column_data[0]==None: # if header of the description table is NULL, skip that row.
            r+=1
            continue            
        else:
            description_data.append(column_data)
            j+=1
            print(j)
            print(column_data)
            r+=1
            print(r)

    for d in description_data:
        pdf.set_font(family="Arial",style='B',size=10)
        pdf.cell(w=50,h=8,txt=str(d[0]),border=1,align='C')
        if len(d)>1: # i.e description table has value also apart from the header.
            pdf.cell(w=70,h=8,txt=str(d[1]),border=1,align='C',ln=1)
        else:
            pdf.cell(w=70,h=8,txt='',border=1,align='C',ln=1)

    add_headers(pdf,headerPos=90)
    ####### for actual data-table ######

    #startYPos = 26
    startYPos=106


    ### try to find the beginning position of the main table.
    for i in range(18,25):
        if ws['B'+str(i)].value=='Audit Parameters' or ws['C'+str(i)].value=='Question Type':
            begin = i+1
            break
        else:
            continue




    dataList = []


    for i in range(begin,k+1):#range(19,k+1):    
        
        obs = [str(i)]+[ws[a+str(i)].value for a in columns]
        print(obs)
        dataList.append(obs)

    dataList=sorted(dataList,key=lambda x:(x[-1] is None,x[0])) 


    cellHeight=80
    for obs in dataList:
        if (270-startYPos<80):
            print('max height: ',startYPos)
            print('adding new page')
            pdf.add_page()
            add_headers(pdf,headerPos=10)
            #global startYPos
            startYPos=26
   
        startX = 10

        for z in range(len(obs)):
        
            pdf.set_xy(xCoords[z],startYPos)
            if z in [0,1,2,3,4,5,6,7,8]:
                text = str(obs[z])
                if '\n' in text:
                    text=text.replace('\n','. ')
                pdf.multi_cell(w=widths[z],h=8,txt=text,align='C')
                pdf.line(xCoords[z], startYPos, xCoords[z]+widths[z],startYPos) # a line above the cell at: x1,y1,x2,y2 
                pdf.line(xCoords[z], startYPos+cellHeight, xCoords[z]+widths[z],startYPos+cellHeight) # a line below the cell at: x1,y1,x2,y2
                pdf.line(xCoords[z],startYPos,xCoords[z],startYPos+cellHeight) # a line on the left of the cell
                pdf.line(xCoords[z]+widths[z],startYPos,xCoords[z]+widths[z],startYPos+cellHeight) # a line on the right of the cell

            if z==9 and obs[z]!=None:
                text = obs[z]
                print('URL:',text)
                imageName = text.split('/')[-1]
                imagePath = imageFolder+imageName
                if os.path.exists(imagePath)==False:
                    imageFile = open(imagePath,'wb')
                    print('downloading from URL: ',obs[z])
                    imageRes=requests.get(obs[z])
                    if imageRes.status_code==200:
                        for chunk in imageRes.iter_content(100000):
                            imageFile.write(chunk)
                        imageFile.close()
                    print('image downloaded and saved')
                else:
                    print('image already exists')
                print('attaching image at x,y:',(xCoords[z],startYPos))
                try:
                    pdf.image(imagePath,x=xCoords[z],y=startYPos,w=70,h=80)
                    pdf.link(x=xCoords[z],y=startYPos,w=70,h=80,link=text)
                except Exception as e:
                    pdf.multi_cell(w=70,h=8,txt='Broken link: '+text,align='C')
                    pdf.link(x=xCoords[z],y=startYPos,w=70,h=80,link=text)

            pdf.line(xCoords[z], startYPos, xCoords[z]+widths[z],startYPos) # a line above the cell at: x1,y1,x2,y2 
            pdf.line(xCoords[z], startYPos+cellHeight, xCoords[z]+widths[z],startYPos+cellHeight) # a line below the cell at: x1,y1,x2,y2
            pdf.line(xCoords[z],startYPos,xCoords[z],startYPos+cellHeight) # a line on the left of the cell
            pdf.line(xCoords[z]+widths[z],startYPos,xCoords[z]+widths[z],startYPos+cellHeight) # a line on the right of the cell
                
        
                
                
        
        startYPos+=cellHeight
        print('reached Ypos height:',startYPos)   
        print(obs)

    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y %Hh%Mm%Ss")
    pdfName = 'newConvertedFile_'+dt_string+'.pdf'.replace('/','-')
    print(pdfName)
    pdf.output(pdfName)
    return pdfName


# excel2pdf('C:\\Users\\Joel\\NOCPL\\BranchReport_pdfs\\Copy of Individual Branch Report (2).xlsx')


   

